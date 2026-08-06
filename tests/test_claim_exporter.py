from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from material_claim_manager.exporter import export_rows
from material_claim_manager.models import LedgerRow
from material_claim_manager.storage import ClaimOverrideStore


def test_export_uses_official_supply_and_does_not_modify_source(tmp_path: Path) -> None:
    source = tmp_path / "ledger.xlsx"
    source.write_bytes(b"source workbook bytes")
    before = source.read_bytes()
    row = LedgerRow(
        source_row=8,
        intake_date=date(2026, 6, 8),
        trade="잡자재",
        item="고급천막",
        spec="10*10",
        unit="장",
        length="",
        quantity=5,
        unit_price=70_000,
        amount=349_999,
        vendor="A상사",
        usage="휴게실",
        note="",
        source_sheet="잡자재",
    )
    store = ClaimOverrideStore(str(source))
    output = tmp_path / "claim.xlsx"
    export_rows(str(output), [row], store, site_name="테스트 현장")
    assert source.read_bytes() == before
    workbook = load_workbook(output, data_only=False)
    sheet = workbook["청구대상"]
    assert sheet.cell(5, 11).value == 349_999
    assert sheet.cell(5, 12).value == 350_000
    assert sheet.cell(5, 13).value == -1
    assert sheet.cell(5, 18).value == 35_000
    assert sheet.cell(5, 19).value == 384_999
    workbook.close()
