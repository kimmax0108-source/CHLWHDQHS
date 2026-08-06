from datetime import datetime
from pathlib import Path
import sys

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from material_claim_manager.excel_io import load_ledger


def test_loader_prefers_second_named_ledger_sheet(tmp_path):
    path = tmp_path / "현장대장_2026.xlsx"
    workbook = Workbook()
    summary = workbook.active
    summary.title = "주자재총괄표"
    summary["A1"] = "총괄표"
    ledger = workbook.create_sheet("자재입출고대장")
    ledger.append(["자재입출고 대장"])
    ledger.append([])
    ledger.append(["현장명 : 테스트 현장"])
    ledger.append([])
    ledger.append(["일자", "공종", "품명", "규격", "단위", "길이", "", "입고", "", "구입처", "용도", "비고"])
    ledger.append(["", "", "", "", "", "", "수량", "단가", "금액", "", "", ""])
    ledger.append([])
    ledger.append([datetime(2026, 6, 1), "잡자재", "고급천막", "10*10", "장", "", 5, 70000, 350000, "삼성건철", "휴게실", ""])
    workbook.save(path)

    data = load_ledger(str(path))
    assert data.sheet_name == "자재입출고대장"
    assert data.sheet_names == ["주자재총괄표", "자재입출고대장"]
    assert data.site_name == "테스트 현장"
    assert len(data.rows) == 1
    assert data.rows[0].item == "고급천막"


def test_standard_workbook_loads_material_sheets_with_header_mapping(tmp_path):
    path = tmp_path / "표준_자재입출고대장_2026.xlsx"
    workbook = Workbook()
    guide = workbook.active
    guide.title = "사용안내"
    guide["A1"] = "안내"

    misc = workbook.create_sheet("잡자재")
    misc.append(["자재 입(출)고 대장 [ 잡자재 ]"])
    misc.append(["현장명 : 표준 테스트 현장"])
    misc.append([
        "일자", "공종", "공종명", "품명", "규격", "단위", "길이", "수량",
        "단가", "공급가액", "부가세", "계", "구입처", "용도", "비고",
        "공무", "관리", "소장",
    ])
    misc.append([
        datetime(2026, 6, 1), "잡자재", "", "고급천막", "10*10", "장", "", 5,
        70000, 350000, 35000, 385000, "A상사", "휴게실", "", "", "", "",
    ])

    main = workbook.create_sheet("주자재")
    main.append(["자재 입(출)고 대장 [ 주자재 ]"])
    main.append(["현장명 : 표준 테스트 현장"])
    main.append([
        "일자", "공종", "품명", "규격", "단위", "길이", "수량", "단가",
        "공급가액", "부가세", "계", "구입처", "용도", "비고(사용위치)",
        "공무", "관리", "소장",
    ])
    main.append([
        datetime(2026, 6, 2), "주자재", "레미콘", "25-21-150", "㎥", "", 10,
        90000, 900000, 90000, 990000, "B레미콘", "타설용", "101동", "", "", "",
    ])

    safety = workbook.create_sheet("안전")
    safety.append(["자재 입(출)고 대장 [ 안전 ]"])
    safety.append(["현장명 : 표준 테스트 현장"])
    safety.append([
        "일자", "공종", "품명", "규격", "단위", "길이", "수량", "단가",
        "공급가액", "부가세", "계", "구입처", "용도", "비고", "공무", "관리", "소장",
    ])
    safety.append([
        datetime(2026, 6, 3), "안전", "안전모", "백색", "개", "", 20,
        6000, 120000, 12000, 132000, "C안전", "직원용", "", "", "", "",
    ])
    workbook.save(path)

    from material_claim_manager.excel_io import detect_ledger_sheets, load_ledgers

    assert detect_ledger_sheets(str(path)) == ["잡자재", "주자재", "안전"]
    data = load_ledgers(str(path))
    assert data.loaded_sheet_names == ["잡자재", "주자재", "안전"]
    assert data.site_name == "표준 테스트 현장"
    assert len(data.rows) == 3
    assert [row.source_sheet for row in data.rows] == ["잡자재", "주자재", "안전"]
    assert data.rows[0].item == "고급천막"
    assert data.rows[0].quantity == 5
    assert data.rows[1].vendor == "B레미콘"
