from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from .models import LedgerRow
from .services import (
    effective_claim_month,
    effective_trade,
    processing_status,
    rollover_status,
)
from .storage import ClaimOverrideStore


HEADERS = [
    "원본 시트",
    "일자",
    "원본 공종",
    "청구분류",
    "품명",
    "규격",
    "단위",
    "길이",
    "입고수량",
    "단가",
    "원본 공급가액",
    "계산금액(수량×단가)",
    "차이금액",
    "검토상태",
    "구입처",
    "용도",
    "비고",
    "부가세(10%)",
    "최종금액",
    "청구년월",
    "이월상태",
    "처리상태",
    "관리메모",
]


def export_rows(
    output_path: str,
    rows: Iterable[LedgerRow],
    store: ClaimOverrideStore,
    *,
    site_name: str = "",
    title: str = "자재 입고 청구대상",
) -> None:
    values = list(rows)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "청구대상"

    last_column = get_column_letter(len(HEADERS))
    worksheet.merge_cells(f"A1:{last_column}1")
    worksheet["A1"] = title
    worksheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    worksheet["A1"].fill = PatternFill("solid", fgColor="174EA6")
    worksheet["A1"].alignment = Alignment(horizontal="center")
    worksheet["A2"] = f"현장명: {site_name or '-'}"
    worksheet.merge_cells(f"A2:{last_column}2")

    worksheet.append([])
    worksheet.append(HEADERS)
    header_row = 4
    for cell in worksheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="20324A")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    mismatch_fill = PatternFill("solid", fgColor="FDE2E2")
    for row in values:
        worksheet.append(
            [
                row.source_sheet or "단일대장",
                row.intake_date,
                row.trade,
                effective_trade(row, store),
                row.item,
                row.spec,
                row.unit,
                row.length,
                row.quantity,
                row.unit_price if row.unit_price_entered else None,
                row.amount,
                row.calculated_amount,
                row.amount_difference,
                row.amount_review_status,
                row.vendor,
                row.usage,
                row.note,
                row.vat,
                row.total,
                effective_claim_month(row, store),
                rollover_status(row, store),
                processing_status(row, store),
                store.management_note(row.fingerprint),
            ]
        )
        if row.amount_review_status == "불일치":
            for cell in worksheet[worksheet.max_row]:
                cell.fill = mismatch_fill

    last_row = worksheet.max_row
    if last_row >= 5:
        table = Table(displayName="ClaimRowsTable", ref=f"A4:{last_column}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        worksheet.add_table(table)

    worksheet.freeze_panes = "A5"
    worksheet.sheet_view.showGridLines = False
    widths = [
        13, 12, 11, 11, 18, 20, 9, 9, 12, 14, 18, 20, 14, 12, 20, 38, 18,
        16, 17, 13, 13, 13, 25,
    ]
    for index, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for row_index in range(5, last_row + 1):
        worksheet.cell(row_index, 2).number_format = "yy.mm.dd"
        worksheet.cell(row_index, 9).number_format = "#,##0.###"
        for column in (10, 11, 12, 13, 18, 19):
            worksheet.cell(row_index, column).number_format = "#,##0"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
