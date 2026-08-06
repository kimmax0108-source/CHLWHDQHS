from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional
import re

from .models import LedgerData, LedgerRow

STANDARD_LEDGER_SHEETS = ("잡자재", "주자재", "안전")


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number(value: Any) -> tuple[float, bool]:
    if value in (None, ""):
        return 0.0, False
    if isinstance(value, (int, float)):
        return float(value), True
    cleaned = str(value).strip().replace(",", "").replace("원", "")
    try:
        return (float(cleaned), True) if cleaned else (0.0, False)
    except ValueError:
        return 0.0, False


def _infer_base_year(path: str, fallback: int) -> int:
    candidates = re.findall(r"(?<!\d)(20\d{2})(?!\d)", Path(path).name)
    return int(candidates[-1]) if candidates else fallback


def parse_date(value: Any, base_year: int, datemode: int = 0) -> Optional[date]:
    """'01.02, '26.01.02, 2026-01-02, 엑셀 날짜값을 모두 인식한다."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)) and value > 1000:
        try:
            import xlrd

            return xlrd.xldate_as_datetime(float(value), datemode).date()
        except Exception:
            try:
                from openpyxl.utils.datetime import from_excel

                converted = from_excel(float(value))
                return converted.date() if isinstance(converted, datetime) else converted
            except Exception:
                return None

    raw = (
        str(value)
        .strip()
        .lstrip("'")
        .replace("년", ".")
        .replace("월", ".")
        .replace("일", "")
    )
    raw = re.sub(r"[\-/]", ".", raw)
    raw = re.sub(r"\.+", ".", raw).strip(".")
    parts = [p for p in raw.split(".") if p]
    try:
        if len(parts) == 2:
            year, month, day = base_year, int(parts[0]), int(parts[1])
        elif len(parts) == 3:
            year_value = int(parts[0])
            year = 2000 + year_value if year_value < 100 else year_value
            month, day = int(parts[1]), int(parts[2])
        else:
            return None
        return date(year, month, day)
    except (TypeError, ValueError):
        return None


def _normalize_header(value: Any) -> str:
    return (
        re.sub(r"\s+", "", _text(value))
        .replace("(", "")
        .replace(")", "")
        .replace("[", "")
        .replace("]", "")
    )


def _normalize_sheet_name(value: str) -> str:
    return re.sub(r"\s+", "", value).casefold()


def _choose_sheet_name(names: list[str], requested: str | None = None) -> str:
    if not names:
        raise ValueError("통합문서에 시트가 없습니다.")
    if requested and requested in names:
        return requested
    target = _normalize_sheet_name("자재입출고대장")
    for name in names:
        if _normalize_sheet_name(name) == target:
            return name
    return names[1] if len(names) > 1 else names[0]


def preferred_ledger_sheets(names: list[str]) -> list[str]:
    """표준대장의 잡자재·주자재·안전 시트를 우선 반환하고, 없으면 기존 대장 1개를 반환한다."""
    normalized = {_normalize_sheet_name(name): name for name in names}
    standard = [
        normalized[_normalize_sheet_name(target)]
        for target in STANDARD_LEDGER_SHEETS
        if _normalize_sheet_name(target) in normalized
    ]
    if standard:
        return standard
    return [_choose_sheet_name(names)]


def _find_header(matrix: list[list[Any]]) -> int:
    for row_index, row in enumerate(matrix[:50]):
        normalized = [_normalize_header(value) for value in row[:20]]
        if any(value in {"일자", "입고일"} for value in normalized) and any(
            value == "공종" for value in normalized
        ):
            return row_index
    raise ValueError("일자·공종 헤더를 찾지 못했습니다.")


FIELD_ALIASES: dict[str, set[str]] = {
    "date": {"일자", "입고일"},
    "trade": {"공종"},
    "item": {"품명"},
    "spec": {"규격"},
    "unit": {"단위"},
    "length": {"길이"},
    "quantity": {"수량", "입고수량", "수량입고"},
    "unit_price": {"단가"},
    "amount": {"공급가액", "금액", "입고금액"},
    "vendor": {"구입처", "업체", "거래처"},
    "usage": {"용도", "사용용도"},
    "note": {"비고", "비고사용위치", "사용위치"},
}


def _header_mapping(matrix: list[list[Any]], header_top: int) -> tuple[dict[str, int], int]:
    top = [_normalize_header(value) for value in matrix[header_top][:30]]
    next_row = (
        [_normalize_header(value) for value in matrix[header_top + 1][:30]]
        if header_top + 1 < len(matrix)
        else []
    )
    mapping: dict[str, int] = {}
    recognized_subheader = False

    for index in range(max(len(top), len(next_row))):
        primary = top[index] if index < len(top) else ""
        secondary = next_row[index] if index < len(next_row) else ""
        candidates = [primary, secondary]
        for field, aliases in FIELD_ALIASES.items():
            if field in mapping:
                continue
            if any(candidate in aliases for candidate in candidates if candidate):
                mapping[field] = index
                if secondary in aliases and primary not in aliases:
                    recognized_subheader = True
                break

    required = {"date", "trade", "item", "quantity", "unit_price", "amount", "vendor"}
    missing = sorted(required.difference(mapping))
    if missing:
        raise ValueError(f"필수 열을 찾지 못했습니다: {', '.join(missing)}")
    return mapping, header_top + (2 if recognized_subheader else 1)


def _extract_site_name(matrix: list[list[Any]]) -> str:
    for row in matrix[:15]:
        for value in row[:15]:
            content = _text(value)
            if "현장명" in content:
                return content.split(":", 1)[-1].strip() if ":" in content else content
    return ""


def _at(row: list[Any], mapping: dict[str, int], field: str) -> Any:
    index = mapping.get(field)
    return row[index] if index is not None and index < len(row) else None


def _rows_from_matrix(
    matrix: list[list[Any]],
    data_start: int,
    base_year: int,
    datemode: int,
    mapping: dict[str, int],
    *,
    source_sheet: str = "",
) -> list[LedgerRow]:
    result: list[LedgerRow] = []
    previous_date: Optional[date] = None

    for index in range(data_start, len(matrix)):
        raw = list(matrix[index])
        date_value = _at(raw, mapping, "date")
        trade = _at(raw, mapping, "trade")
        item = _at(raw, mapping, "item")
        spec = _at(raw, mapping, "spec")
        unit = _at(raw, mapping, "unit")
        length = _at(raw, mapping, "length")
        quantity_value = _at(raw, mapping, "quantity")
        unit_price_value = _at(raw, mapping, "unit_price")
        amount_value = _at(raw, mapping, "amount")
        vendor = _at(raw, mapping, "vendor")
        usage = _at(raw, mapping, "usage")
        note = _at(raw, mapping, "note")

        has_identity = any(_text(value) for value in (trade, item, vendor, usage, note))
        parsed_date = parse_date(date_value, base_year, datemode)
        if parsed_date is None and has_identity:
            parsed_date = previous_date
        if parsed_date is None:
            continue
        previous_date = parsed_date

        quantity, quantity_entered = _number(quantity_value)
        unit_price, unit_price_entered = _number(unit_price_value)
        amount, amount_entered = _number(amount_value)
        if not has_identity and not amount_entered:
            continue

        result.append(
            LedgerRow(
                source_row=index + 1,
                intake_date=parsed_date,
                trade=_text(trade),
                item=_text(item),
                spec=_text(spec),
                unit=_text(unit),
                length=_text(length),
                quantity=quantity,
                unit_price=unit_price,
                amount=amount,
                vendor=_text(vendor),
                usage=_text(usage),
                note=_text(note),
                source_sheet=source_sheet,
                quantity_entered=quantity_entered,
                unit_price_entered=unit_price_entered,
                amount_entered=amount_entered,
            )
        )
    return result


def list_sheet_names(path: str) -> list[str]:
    extension = Path(path).suffix.lower()
    if extension == ".xls":
        import xlrd

        return xlrd.open_workbook(path, on_demand=True).sheet_names()
    if extension in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        names = list(workbook.sheetnames)
        workbook.close()
        return names
    raise ValueError("지원 형식은 .xls, .xlsx, .xlsm 입니다.")


def detect_ledger_sheets(path: str) -> list[str]:
    return preferred_ledger_sheets(list_sheet_names(path))


def load_ledger(
    path: str,
    base_year: int = 2026,
    *,
    sheet_name: str | None = None,
) -> LedgerData:
    """기존 단일 자재입출고대장 호환 로더."""
    extension = Path(path).suffix.lower()
    base_year = _infer_base_year(path, base_year)
    if extension == ".xls":
        return _load_xls_sheets(path, base_year, [sheet_name] if sheet_name else None, legacy=True)
    if extension in {".xlsx", ".xlsm"}:
        return _load_xlsx_sheets(path, base_year, [sheet_name] if sheet_name else None, legacy=True)
    raise ValueError("지원 형식은 .xls, .xlsx, .xlsm 입니다.")


def load_ledgers(
    path: str,
    base_year: int = 2026,
    *,
    sheet_names: list[str] | tuple[str, ...] | None = None,
) -> LedgerData:
    """표준 자재입출고대장의 여러 시트를 한 번에 읽어 통합 데이터로 반환한다."""
    extension = Path(path).suffix.lower()
    base_year = _infer_base_year(path, base_year)
    if extension == ".xls":
        return _load_xls_sheets(path, base_year, list(sheet_names) if sheet_names else None)
    if extension in {".xlsx", ".xlsm"}:
        return _load_xlsx_sheets(path, base_year, list(sheet_names) if sheet_names else None)
    raise ValueError("지원 형식은 .xls, .xlsx, .xlsm 입니다.")


def _load_xls_sheets(
    path: str, base_year: int, requested: list[str] | None, *, legacy: bool = False
) -> LedgerData:
    import xlrd

    book = xlrd.open_workbook(path, formatting_info=False)
    names = book.sheet_names()
    selected = requested or preferred_ledger_sheets(names)
    selected = [name for name in selected if name in names]
    if not selected:
        selected = [_choose_sheet_name(names)]

    all_rows: list[LedgerRow] = []
    site_name = ""
    first_header = 0
    first_data = 0
    multi_standard = len(selected) > 1 or any(
        _normalize_sheet_name(name) in {_normalize_sheet_name(v) for v in STANDARD_LEDGER_SHEETS}
        for name in selected
    )
    for sheet_name in selected:
        sheet = book.sheet_by_name(sheet_name)
        max_columns = min(max(sheet.ncols, 18), 40)
        matrix = [
            [sheet.cell_value(row, column) for column in range(max_columns)]
            for row in range(sheet.nrows)
        ]
        header_top = _find_header(matrix)
        mapping, data_start = _header_mapping(matrix, header_top)
        source_sheet = "" if legacy else sheet_name
        rows = _rows_from_matrix(
            matrix, data_start, base_year, book.datemode, mapping, source_sheet=source_sheet
        )
        all_rows.extend(rows)
        site_name = site_name or _extract_site_name(matrix)
        if not first_header:
            first_header = header_top + 1
            first_data = data_start + 1

    return LedgerData(
        path=path,
        sheet_name=selected[0] if len(selected) == 1 else ", ".join(selected),
        site_name=site_name,
        rows=all_rows,
        header_row=first_header,
        data_start_row=first_data,
        sheet_names=names,
        loaded_sheet_names=selected,
    )


def _load_xlsx_sheets(
    path: str, base_year: int, requested: list[str] | None, *, legacy: bool = False
) -> LedgerData:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    names = list(workbook.sheetnames)
    selected = requested or preferred_ledger_sheets(names)
    selected = [name for name in selected if name in names]
    if not selected:
        selected = [_choose_sheet_name(names)]

    all_rows: list[LedgerRow] = []
    site_name = ""
    first_header = 0
    first_data = 0
    standard_normalized = {_normalize_sheet_name(v) for v in STANDARD_LEDGER_SHEETS}
    multi_standard = len(selected) > 1 or any(
        _normalize_sheet_name(name) in standard_normalized for name in selected
    )

    for sheet_name in selected:
        worksheet = workbook[sheet_name]
        matrix: list[list[Any]] = []
        for raw_row in worksheet.iter_rows(values_only=True):
            row = list(raw_row[:40])
            matrix.append(row)
        header_top = _find_header(matrix)
        mapping, data_start = _header_mapping(matrix, header_top)
        source_sheet = "" if legacy else sheet_name
        rows = _rows_from_matrix(
            matrix, data_start, base_year, 0, mapping, source_sheet=source_sheet
        )
        all_rows.extend(rows)
        site_name = site_name or _extract_site_name(matrix)
        if not first_header:
            first_header = header_top + 1
            first_data = data_start + 1

    workbook.close()
    return LedgerData(
        path=path,
        sheet_name=selected[0] if len(selected) == 1 else ", ".join(selected),
        site_name=site_name,
        rows=all_rows,
        header_row=first_header,
        data_start_row=first_data,
        sheet_names=names,
        loaded_sheet_names=selected,
    )
